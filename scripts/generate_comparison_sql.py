"""Generate the read-only server comparison report.

Reads the source Excel ("data" tab), applies the confirmed mapping.md
column mapping, compares against ita.snowserver (read-only), and writes
four output files, all regenerated together on every run:
  - output/compare_servers.sql   (SELECT-only comparison report, plus concrete
    INSERT/UPDATE recommendations written as SQL comments for human review)
  - output/comparison_report.xlsx (same findings, for human review in Excel):
      * "New Servers" sheet       - servers in source but not in the DB
      * "NULL Field Updates" sheet - DB fields that are NULL and source can fill
      * "Duplicates Resolved" sheet - ServerNames duplicated in the source Excel;
        one distinct row per ServerName (the one with the most recent
        Updated/Sysupdatedon timestamp) is kept and used in the New Servers/
        NULL Field Updates/Matched Comparison sheets, the discarded row is
        listed here for reference. (DB-side duplicates are deduped the same
        way for matching, but not listed here.)
  - output/new_servers.csv        ("New Servers" sheet, as CSV)
  - output/null_field_updates.csv ("NULL Field Updates" sheet, as CSV)

This script never runs INSERT/UPDATE/DELETE against ita.snowserver.
"""
import csv
import os
from collections import OrderedDict, defaultdict
from datetime import datetime
from contextlib import contextmanager

import openpyxl
import pyodbc
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter

load_dotenv()

SOURCE_PATH = r"C:\Feven\Hanna\2026\Gainwell Servers_20260820.xlsx"
SHEET_NAME = "data"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")

DB_SCHEMA = "ita"
DB_TABLE = "snowserver"
KEY_SOURCE_COL = "Name"
KEY_DB_COL = "Servername"

# Confirmed source-column -> ita.snowserver-column mapping (see mapping.md).
# "Discovery source" and "Updates" are intentionally excluded (no reliable
# DB destination column) per user decision on 2026-08-25.
# "Assigned to" is intentionally excluded (Primaryapplicationadmin was an
# incorrect target column) per user decision on 2026-08-28.
MAPPING = OrderedDict(
    [
        ("System name", "System"),
        ("IP Address", "IPaddress"),
        ("Environment", "Environment"),
        ("Vendor", "Vendor"),
        ("Location", "location"),
        ("Purpose", "Purpose"),
        ("Used for", "Usedfor"),
        ("Install Status", "Installstatus"),
        ("Operational status", "Operationalstatus"),
        ("Operating System", "OS"),
        ("OS Version", "OSversion"),
        ("Support group", "Supportgroup"),
        ("Supported by", "Supportedby"),
        ("Host admin  group", "Hostadmin"),  # real header has two spaces
        ("Host admin primary", "Hostadminprimary"),
        ("Managed By Group", "Managedbygroup"),
        ("Managed by", "ManagedBy"),
        ("Owned by", "Ownedby"),
        ("Classification", "Classification"),
        ("Class", "Sysclassname"),
        ("Maintenance schedule", "Maintenanceschedule"),
        ("Most recent discovery", "Lastdiscovered"),
        ("Updated", "Sysupdatedon"),
        ("Updated by", "Sysupdatedby"),
        ("First discovered", "FirstDiscovered"),
    ]
)
DB_COLUMNS = [KEY_DB_COL] + list(MAPPING.values())


@contextmanager
def impersonation_if_configured():
    # Only needed because the direct trusted connection fails with "untrusted domain".
    user_spec = os.environ.get("IMPERSONATE_USER")
    password = os.environ.get("IMPERSONATE_PASSWORD")
    if not user_spec or not password:
        yield
        return
    import win32con
    import win32security

    domain, username = user_spec.split("\\", 1)
    handle = win32security.LogonUser(
        username, domain, password,
        win32con.LOGON32_LOGON_NEW_CREDENTIALS, win32con.LOGON32_PROVIDER_WINNT50,
    )
    win32security.ImpersonateLoggedOnUser(handle)
    try:
        yield
    finally:
        win32security.RevertToSelf()
        handle.Close()


def load_source_records():
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header)}

    missing = [c for c in [KEY_SOURCE_COL] + list(MAPPING.keys()) if c not in idx]
    if missing:
        raise SystemExit(f"Source sheet is missing expected column(s): {missing}")

    by_name = defaultdict(list)
    total_rows = 0
    for row in rows:
        name = row[idx[KEY_SOURCE_COL]]
        if name is None or str(name).strip() == "":
            continue
        total_rows += 1
        name = str(name).strip()
        record = {KEY_DB_COL: name}
        for src_col, db_col in MAPPING.items():
            val = row[idx[src_col]]
            if isinstance(val, str):
                val = val.strip() or None
            record[db_col] = val
        by_name[norm_key(name)].append(record)
    return by_name, total_rows


def fetch_db_records(conn):
    cur = conn.cursor()
    cols_sql = ", ".join(f"[{c}]" for c in DB_COLUMNS)
    cur.execute(f"SELECT {cols_sql} FROM {DB_SCHEMA}.{DB_TABLE}")
    by_name = defaultdict(list)
    for row in cur.fetchall():
        rec = dict(zip(DB_COLUMNS, row))
        name = rec[KEY_DB_COL]
        if name is None or str(name).strip() == "":
            continue
        by_name[norm_key(str(name))].append(rec)
    return by_name


def sql_literal(val):
    if val is None:
        return "NULL"
    return "N'" + str(val).replace("'", "''") + "'"


def norm_key(name):
    # ita.snowserver.Servername collation is SQL_Latin1_General_CP1_CI_AS (case-insensitive);
    # group/match on uppercase so Python agrees with what SQL Server considers equal.
    return name.strip().upper()


def pick_latest(records, date_field="Sysupdatedon"):
    # When a ServerName has multiple rows, keep the single distinct row with the
    # most recent Updated/Sysupdatedon timestamp; rows with no timestamp sort oldest.
    def sort_key(rec):
        val = rec.get(date_field)
        return val if isinstance(val, datetime) else datetime.min

    return max(records, key=sort_key)


def autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 60)


def write_excel_report(path, new_servers, null_field_updates, resolved_duplicates, matched_details):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "New Servers"
    ws.append(["ServerName"] + list(MAPPING.values()))
    for rec in new_servers:
        ws.append([rec[KEY_DB_COL]] + [rec.get(c) for c in MAPPING.values()])
    autosize(ws)

    ws = wb.create_sheet("NULL Field Updates")
    ws.append(["ServerName", "FieldName", "CurrentValue", "ProposedValue"])
    for name, field, current, proposed in null_field_updates:
        ws.append([name, field, current, proposed])
    autosize(ws)

    ws = wb.create_sheet("Duplicates Resolved")
    ws.append(["ServerName", "Source", "KeptForComparison", "Reason"] + DB_COLUMNS)
    for origin, name, rec, kept in resolved_duplicates:
        reason = (
            f"Duplicate ServerName in {origin}; kept the row with the most recent "
            "Updated timestamp for comparison" if kept else
            f"Duplicate ServerName in {origin}; discarded in favor of a more recently updated row"
        )
        ws.append([name, origin, kept, reason] + [rec.get(c) for c in DB_COLUMNS])
    autosize(ws)

    ws = wb.create_sheet("Matched Comparison")
    header = ["ServerName"]
    for src_col, db_col in MAPPING.items():
        header += [f"{src_col} (source)", f"{db_col} (DB)"]
    ws.append(header)
    for name, src, db_rec in matched_details:
        row = [name]
        for db_col in MAPPING.values():
            row += [src.get(db_col), db_rec.get(db_col)]
        ws.append(row)
    autosize(ws)

    wb.save(path)


def write_csv_reports(new_servers_path, null_updates_path, new_servers, null_field_updates):
    with open(new_servers_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ServerName"] + list(MAPPING.values()))
        for rec in new_servers:
            writer.writerow([rec[KEY_DB_COL]] + [rec.get(c) for c in MAPPING.values()])

    with open(null_updates_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ServerName", "FieldName", "CurrentValue", "ProposedValue"])
        for name, field, current, proposed in null_field_updates:
            writer.writerow([name, field, current, proposed])


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    source_by_name, total_source_rows = load_source_records()

    conn_str = os.environ["DB_CONN"]
    with impersonation_if_configured():
        conn = pyodbc.connect(conn_str)
    try:
        db_by_name = fetch_db_records(conn)
    finally:
        conn.close()

    resolved_duplicates = []  # rows for the "Duplicates Resolved" report - source-side only, per user request
    clean_source = {}  # normalized name -> single distinct source record (original casing kept in the record)
    for name, records in source_by_name.items():
        if len(records) > 1:
            # ServerName isn't unique here; keep one distinct row (most recently updated) for comparison.
            chosen = pick_latest(records)
            for rec in records:
                resolved_duplicates.append(("source", rec[KEY_DB_COL], rec, rec is chosen))
            clean_source[name] = chosen
        else:
            clean_source[name] = records[0]

    # DB-side duplicates are still deduped (most recent row wins) for matching, but not
    # listed in the report - only source-file duplicates are relevant for human review here.

    usable_names = list(clean_source.keys())  # every ServerName is now comparable (duplicates resolved above)

    new_servers = []
    matched_servers = []
    matched_details = []  # (ServerName, source record, db record) for the side-by-side comparison sheet
    null_field_updates = []
    for name in usable_names:
        src = clean_source[name]
        db_matches = db_by_name.get(name)
        if not db_matches:
            new_servers.append(src)
            continue
        matched_servers.append(name)
        db_rec = pick_latest(db_matches)  # keep the single distinct (most recently updated) DB row
        matched_details.append((src[KEY_DB_COL], src, db_rec))
        for db_col in MAPPING.values():
            db_val = db_rec.get(db_col)
            src_val = src.get(db_col)
            # Only flag NULL DB fields the source can fill; never touch already-populated fields.
            if (db_val is None or str(db_val).strip() == "") and src_val not in (None, ""):
                null_field_updates.append((src[KEY_DB_COL], db_col, db_val, src_val))

    # --- output/comparison_report.xlsx ---
    xlsx_path = os.path.join(OUTPUT_DIR, "comparison_report.xlsx")
    write_excel_report(xlsx_path, new_servers, null_field_updates, resolved_duplicates, matched_details)

    # --- output/new_servers.csv & output/null_field_updates.csv ---
    new_servers_csv = os.path.join(OUTPUT_DIR, "new_servers.csv")
    null_updates_csv = os.path.join(OUTPUT_DIR, "null_field_updates.csv")
    write_csv_reports(new_servers_csv, null_updates_csv, new_servers, null_field_updates)

    # --- output/compare_servers.sql ---
    cols = [KEY_DB_COL] + list(MAPPING.values())
    sql_path = os.path.join(OUTPUT_DIR, "compare_servers.sql")
    with open(sql_path, "w", encoding="utf-8") as f:
        f.write("-- Auto-generated by scripts/generate_comparison_sql.py\n")
        f.write("-- Read-only recommendation report comparing the source Excel against\n")
        f.write(f"-- {DB_SCHEMA}.{DB_TABLE}. All comparison logic already ran in Python; this file\n")
        f.write("-- only lists the results as commented-out INSERT/UPDATE statements below.\n")
        f.write("-- ServerNames duplicated in the source Excel are resolved, not excluded: the\n")
        f.write("-- single distinct row with the most recent Updated timestamp is kept for\n")
        f.write("-- comparison; see the 'Duplicates Resolved' sheet in output/comparison_report.xlsx.\n")
        f.write("-- (DB-side ServerName duplicates are deduped the same way but not listed there.)\n")
        f.write("-- This file contains no live SQL statements - every INSERT/UPDATE line is\n")
        f.write(f"-- commented out. It never INSERTs, UPDATEs, or DELETEs any row in {DB_SCHEMA}.{DB_TABLE}.\n\n")

        f.write("-- Summary\n")
        f.write(f"--   Total source rows (Excel, non-blank ServerName):  {total_source_rows}\n")
        f.write(f"--   Matched servers (found in both source and DB): {len(matched_servers)}\n")
        f.write(f"--   New servers (in source, not in DB):             {len(new_servers)}\n")
        f.write(f"--   NULL-field update recommendations:              {len(null_field_updates)}\n")
        f.write(f"--   Duplicate ServerNames in source resolved (kept the latest Updated row): {len(resolved_duplicates)}\n\n")

        # 3) & 4): concrete INSERT/UPDATE recommendations, commented out so the file
        # can never be run as-is; a human must deliberately uncomment a statement to apply it.
        f.write("-- 1) RECOMMENDED INSERTs for missing servers (commented out - review before uncommenting)\n")
        insert_cols = ", ".join(f"[{c}]" for c in cols)
        for rec in new_servers:
            values = ", ".join(sql_literal(rec.get(c)) for c in cols)
            f.write(
                f"-- INSERT INTO {DB_SCHEMA}.{DB_TABLE} ({insert_cols}) VALUES ({values});\n"
            )
        f.write("\n")

        f.write("-- 2) RECOMMENDED UPDATEs for NULL fields (commented out - review before uncommenting)\n")
        for name, field, _current, proposed in null_field_updates:
            f.write(
                f"-- UPDATE {DB_SCHEMA}.{DB_TABLE} SET [{field}] = {sql_literal(proposed)} "
                f"WHERE {KEY_DB_COL} = {sql_literal(name)} AND [{field}] IS NULL;\n"
            )
        f.write("\n")

    print(f"Matched servers (found in both source and DB): {len(matched_servers)}")
    print(f"New servers: {len(new_servers)}")
    print(f"NULL-field update recommendations: {len(null_field_updates)}")
    print(f"Duplicate ServerNames in source resolved (kept the latest Updated row): {len(resolved_duplicates)}")
    print(f"Total source rows: {total_source_rows}")
    print(f"Wrote: {sql_path}")
    print(f"Wrote: {xlsx_path}")
    print(f"Wrote: {new_servers_csv}")
    print(f"Wrote: {null_updates_csv}")


if __name__ == "__main__":
    main()
