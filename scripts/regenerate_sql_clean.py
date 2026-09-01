#!/usr/bin/env python3
"""
Regenerate sts_server_comparison.sql with proper formatting (7 servers per line)
"""
import openpyxl

# Read source Excel
excel_file = r"C:\Feven\Hanna\2026\Gainwell Servers_20260820.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb['data']

# Find the "Name" column (header in row 1)
name_col_index = None
for cell in ws[1]:  # First row contains headers
    if cell.value and str(cell.value).strip().lower() == 'name':
        name_col_index = cell.column - 1  # Convert to 0-based index
        break

if name_col_index is None:
    print("✗ Error: 'Name' column not found")
    exit(1)

# Extract distinct ServerNames from Excel
excel_servers = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    servername = row[name_col_index]
    if servername and str(servername).strip():
        excel_servers.add(str(servername).strip())

print(f"Extracted {len(excel_servers)} distinct servers")

# Create SQL file with proper formatting
output_file = r"c:\projects\GainwellServers_20260820\output\sts_server_comparison.sql"
servers_list = sorted(list(excel_servers))

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("-- Comparison of source Excel ServerNames against sts.server table\n")
    f.write("-- Generated: 2026-08-26\n")
    f.write("-- Source Excel: Gainwell Servers_20260820.xlsx (data tab, Name column)\n")
    f.write("-- Target Table: sts.server (servername column)\n\n")
    
    f.write("-- Summary\n")
    f.write(f"--   Distinct ServerNames in Excel: {len(excel_servers)}\n\n")
    
    f.write("-- SQL query to check Excel servers against sts.server:\n")
    f.write("-- SELECT DISTINCT Servername FROM sts.server WHERE Servername IN (\n")
    
    # Write 7 servers per line
    for i in range(0, len(servers_list), 7):
        group = servers_list[i:i+7]
        line = "--   " + ", ".join([f"N'{server}'" for server in group])
        if i + 7 < len(servers_list):  # Not the last group
            line += ","
        f.write(line + "\n")
    
    f.write("-- )\n")
    f.write("-- ORDER BY Servername;\n\n")
    
    f.write("-- Distinct ServerNames from Excel (for reference):\n")
    for server in servers_list:
        f.write(f"--   {server}\n")

print(f"✓ File created: {output_file}")
print(f"✓ Formatted with 7 servers per line")
