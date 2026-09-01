#!/usr/bin/env python3
"""
Compare distinct ServerNames from Excel against sts.server table using ita.fnServerName function
"""
import openpyxl

# Read source Excel
excel_file = r"C:\Feven\Hanna\2026\Gainwell Servers_20260820.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb['data']

# Find the "Name" column (header in row 1)
name_col_index = None
for cell in ws[1]:  # First row contains headers
    if cell.value and str(cell.value).strip().lower() == 'name':
        name_col_index = cell.column - 1  # Convert to 0-based index
        break

if name_col_index is None:
    print("✗ Error: 'Name' column not found in Excel header row")
    exit(1)

print(f"Found 'Name' column at index {name_col_index}")

# Extract distinct ServerNames from Excel (Name column)
excel_servers = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    servername = row[name_col_index]  # Extract from Name column
    if servername and str(servername).strip():
        servername_clean = str(servername).strip()
        # Exclude "(unknown)" entries
        if servername_clean.lower() != "(unknown)":
            excel_servers.add(servername_clean)

print(f"Distinct ServerNames from Excel (excluding unknown): {len(excel_servers)}")

# Create SQL file
output_file = r"c:\projects\GainwellServers_20260820\output\sts_server_comparison.sql"
sorted_servers = sorted(list(excel_servers))

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("-- Comparison: Excel ServerNames vs sts.server table\n")
    f.write("-- Generated: 2026-08-26\n")
    f.write("-- Source Excel: Gainwell Servers_20260820.xlsx (data tab, Name column)\n")
    f.write("-- Target Table: sts.server\n")
    f.write(f"-- Total servers from Excel: {len(excel_servers)}\n")
    f.write("-- Function used: ita.fnServerName('Servername', N'....')\n\n")
    
    f.write("-- ============================================================================\n")
    f.write("-- QUERY 1: Count servers from Excel that EXIST in sts.server\n")
    f.write("-- ============================================================================\n")
    f.write("SELECT COUNT(DISTINCT s.Servername) AS 'Servers Found in sts.server'\n")
    f.write("FROM sts.server s\n")
    f.write("WHERE LOWER(s.Servername) IN (\n")
    
    for i, server in enumerate(sorted_servers):
        comma = "," if i < len(sorted_servers) - 1 else ""
        if i % 20 == 0:
            f.write(f"  N'{server}'{comma}")
        else:
            f.write(f" N'{server}'{comma}")
        if (i + 1) % 20 == 0:
            f.write("\n")
    
    f.write("\n);\n\n")
    
    f.write("-- ============================================================================\n")
    f.write("-- QUERY 2: List servers from Excel that EXIST in sts.server\n")
    f.write("-- ============================================================================\n")
    f.write("SELECT DISTINCT s.Servername\n")
    f.write("FROM sts.server s\n")
    f.write("WHERE LOWER(s.Servername) IN (\n")
    
    for i, server in enumerate(sorted_servers):
        comma = "," if i < len(sorted_servers) - 1 else ""
        if i % 20 == 0:
            f.write(f"  N'{server}'{comma}")
        else:
            f.write(f" N'{server}'{comma}")
        if (i + 1) % 20 == 0:
            f.write("\n")
    
    f.write("\n)\n")
    f.write("ORDER BY s.Servername;\n\n")
    
    f.write("-- ============================================================================\n")
    f.write("-- QUERY 3: Find missing servers using ita.fnServerName function\n")
    f.write("-- ============================================================================\n")
    f.write("-- This query uses the ita.fnServerName function to look up each Excel server\n")
    f.write("-- Returns servers from Excel that DO NOT exist in sts.server\n\n")
    f.write("WITH excel_servers AS (\n")
    
    for i, server in enumerate(sorted_servers):
        union = " UNION ALL\n" if i < len(sorted_servers) - 1 else "\n"
        f.write(f"  SELECT ita.fnServerName('Servername', N'{server}') AS Servername{union}")
    
    f.write(")\n")
    f.write("SELECT e.Servername FROM excel_servers e\n")
    f.write("WHERE NOT EXISTS (SELECT 1 FROM sts.server s WHERE LOWER(s.Servername) = LOWER(e.Servername))\n")
    f.write("ORDER BY e.Servername;\n")

print(f"✓ Comparison file created: {output_file}")
print(f"✓ Total servers to check: {len(excel_servers)}")
